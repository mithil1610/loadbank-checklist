from flask import Flask, request, jsonify, send_file
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from threading import Lock, Thread
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import gspread
from google.oauth2.service_account import Credentials
import json
from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__)

# Thread lock for file writing
file_lock = Lock()

# Configuration from environment variables
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', '')
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USERNAME = os.environ.get('SMTP_USERNAME', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
GOOGLE_SHEETS_CREDS = os.environ.get('GOOGLE_SHEETS_CREDS', '')
GOOGLE_SHEET_NAME = os.environ.get('GOOGLE_SHEET_NAME', 'Load Bank Checklist Submissions')

SUBMISSIONS_FILE = 'checklist_submissions.xlsx'

def init_submissions_file():
    """Initialize local Excel file"""
    if not os.path.exists(SUBMISSIONS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Submissions"
        
        headers = ['Submission Date', 'Submission Time', 'Serial Number', 'Checklist Date']
        for i in range(1, 72):
            headers.append(f'Q{i}')
            headers.append(f'Q{i}_Remarks')
        
        ws.append(headers)
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        wb.save(SUBMISSIONS_FILE)

def send_email_notification(submission_data, submission_number):
    """Send email notification on form submission with Excel attachment"""
    if not ADMIN_EMAIL or not SMTP_USERNAME or not SMTP_PASSWORD:
        print("Email not configured - skipping notification")
        return False
    
    try:
        msg = MIMEMultipart('mixed')
        msg['Subject'] = f'New Load Bank Checklist Submission #{submission_number}'
        msg['From'] = SMTP_USERNAME
        msg['To'] = ADMIN_EMAIL
        
        # Create email body
        html = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; background: #f7f7f7;">
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px 10px 0 0;">
                        <h1 style="margin: 0; font-size: 24px;">New Checklist Submission</h1>
                        <p style="margin: 10px 0 0 0; opacity: 0.9;">Submission #{submission_number}</p>
                    </div>
                    
                    <div style="background: white; padding: 30px; border-radius: 0 0 10px 10px;">
                        <h2 style="color: #667eea; margin-top: 0;">Submission Details</h2>
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr style="background: #f7f7f7;">
                                <td style="padding: 12px; border: 1px solid #ddd; font-weight: bold;">Serial Number:</td>
                                <td style="padding: 12px; border: 1px solid #ddd;">{submission_data.get('serialNumber', 'N/A')}</td>
                            </tr>
                            <tr>
                                <td style="padding: 12px; border: 1px solid #ddd; font-weight: bold;">Date:</td>
                                <td style="padding: 12px; border: 1px solid #ddd;">{submission_data.get('date', 'N/A')}</td>
                            </tr>
                            <tr style="background: #f7f7f7;">
                                <td style="padding: 12px; border: 1px solid #ddd; font-weight: bold;">Submitted:</td>
                                <td style="padding: 12px; border: 1px solid #ddd;">{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</td>
                            </tr>
                        </table>
                        
                        <h3 style="color: #667eea; margin-top: 25px;">Summary</h3>
                        <p style="background: #e8f4f8; padding: 15px; border-radius: 5px; border-left: 4px solid #667eea;">
                            A new load bank inspection checklist has been submitted. 
                            The complete data has been saved to your Google Sheet and the Excel file is attached to this email.
                        </p>
                        
                        <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin-top: 20px; border-radius: 5px;">
                            <strong>📎 Attachment:</strong> The updated Excel file with all submissions is attached below.
                        </div>
                        
                        <div style="margin-top: 25px; text-align: center;">
                            <p style="color: #666; font-size: 12px;">
                                This is an automated notification from your Load Bank Checklist System
                            </p>
                        </div>
                    </div>
                </div>
            </body>
        </html>
        """
        
        # Attach HTML body
        msg.attach(MIMEText(html, 'html'))
        
        # Attach Excel file
        if os.path.exists(SUBMISSIONS_FILE):
            try:
                with open(SUBMISSIONS_FILE, 'rb') as f:
                    part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    
                    filename = f'checklist_submissions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                    part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                    msg.attach(part)
                    
                    print(f"✓ Excel file attached: {filename}")
            except Exception as attach_error:
                print(f"✗ Could not attach Excel file: {str(attach_error)}")
        
        # Send email
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(msg)
        
        print(f"✓ Email notification sent to {ADMIN_EMAIL} with attachment")
        return True
        
    except Exception as e:
        print(f"✗ Email error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def save_to_google_sheets(submission_data, submission_number):
    """Save submission to Google Sheets (sheet must exist and be shared with service account)"""
    if not GOOGLE_SHEETS_CREDS:
        print("Google Sheets not configured - skipping")
        return False
    
    try:
        # Parse credentials from environment variable
        creds_dict = json.loads(GOOGLE_SHEETS_CREDS)
        
        # Set up credentials
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(creds)
        
        # Open existing spreadsheet (must be manually created and shared)
        try:
            spreadsheet = client.open(GOOGLE_SHEET_NAME)
            print(f"✓ Opened spreadsheet: {GOOGLE_SHEET_NAME}")
        except gspread.SpreadsheetNotFound:
            service_account_email = creds_dict.get('client_email', 'unknown')
            print(f"✗ ERROR: Spreadsheet '{GOOGLE_SHEET_NAME}' not found!")
            print(f"   Please create it in Google Drive and share with: {service_account_email}")
            return False
        except Exception as open_error:
            print(f"✗ Error opening spreadsheet: {str(open_error)}")
            return False
        
        # Get first worksheet
        try:
            worksheet = spreadsheet.sheet1
        except Exception as ws_error:
            print(f"✗ Error accessing worksheet: {str(ws_error)}")
            return False
        
        # Check if headers exist and add them if needed
        try:
            existing_headers = worksheet.row_values(1) if worksheet.row_count > 0 else []
            
            if not existing_headers:
                headers = ['Submission #', 'Submission Date', 'Submission Time', 'Serial Number', 'Checklist Date']
                for i in range(1, 72):
                    headers.append(f'Q{i}')
                    headers.append(f'Q{i}_Remarks')
                worksheet.append_row(headers)
                print("✓ Headers added to spreadsheet")
                
                # Format header row
                try:
                    worksheet.format('A1:EW1', {
                        'backgroundColor': {'red': 0.4, 'green': 0.49, 'blue': 0.92},
                        'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}},
                        'horizontalAlignment': 'CENTER'
                    })
                    print("✓ Header formatting applied")
                except Exception as format_error:
                    print(f"⚠ Could not format headers: {str(format_error)}")
        except Exception as header_error:
            print(f"⚠ Header error: {str(header_error)}")
        
        # Prepare row data
        now = datetime.now()
        row_data = [
            submission_number,
            now.strftime('%Y-%m-%d'),
            now.strftime('%H:%M:%S'),
            submission_data.get('serialNumber', ''),
            submission_data.get('date', '')
        ]
        
        # Add all questions and remarks
        for i in range(1, 72):
            row_data.append(submission_data.get(f'q{i}', ''))
            row_data.append(submission_data.get(f'remarks{i}', ''))
        
        # Append row to spreadsheet
        worksheet.append_row(row_data)
        
        print(f"✓ Data saved to Google Sheets: {GOOGLE_SHEET_NAME} (Submission #{submission_number})")
        return True
        
    except gspread.exceptions.APIError as api_error:
        print(f"✗ Google Sheets API error: {str(api_error)}")
        if hasattr(api_error, 'response'):
            print(f"   Response: {api_error.response}")
        return False
    except json.JSONDecodeError as json_error:
        print(f"✗ Invalid Google Sheets credentials JSON: {str(json_error)}")
        return False
    except Exception as e:
        print(f"✗ Google Sheets error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

@app.route('/')
def index():
    return send_file('checklist_form.html')

@app.route('/submit', methods=['POST'])
def submit_form():
    try:
        data = request.json
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data received'
            }), 400
        
        submission_number = None
        
        # Initialize file once before lock
        init_submissions_file()
        
        # Critical operation: Save to local Excel
        with file_lock:
            try:
                wb = openpyxl.load_workbook(SUBMISSIONS_FILE)
                ws = wb.active
                
                now = datetime.now()
                row_data = [
                    now.strftime('%Y-%m-%d'),
                    now.strftime('%H:%M:%S'),
                    data.get('serialNumber', ''),
                    data.get('date', '')
                ]
                
                for i in range(1, 72):
                    row_data.append(data.get(f'q{i}', ''))
                    row_data.append(data.get(f'remarks{i}', ''))
                
                ws.append(row_data)
                submission_number = ws.max_row - 1
                wb.save(SUBMISSIONS_FILE)
                
                print(f"✓ Submission #{submission_number} saved to Excel")
                
            except Exception as excel_error:
                print(f"✗ Excel save error: {str(excel_error)}")
                import traceback
                traceback.print_exc()
                raise
        
        # Background tasks - non-blocking
        def background_tasks():
            """Run email and Google Sheets operations in background"""
            try:
                send_email_notification(data, submission_number)
            except Exception as e:
                print(f"Background email error: {str(e)}")
            
            try:
                save_to_google_sheets(data, submission_number)
            except Exception as e:
                print(f"Background Google Sheets error: {str(e)}")
        
        # Start background thread for non-critical operations
        if submission_number:
            thread = Thread(target=background_tasks, daemon=True)
            thread.start()
        
        # Return success immediately
        return jsonify({
            'success': True,
            'message': 'Checklist submitted successfully!',
            'submission_number': submission_number
        }), 200
    
    except Exception as e:
        print(f"✗ Critical error in /submit: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False, 
            'message': f'Submission failed: {str(e)}'
        }), 500

@app.route('/download')
def download_submissions():
    """Download the Excel file with all submissions"""
    if os.path.exists(SUBMISSIONS_FILE):
        return send_file(
            SUBMISSIONS_FILE,
            as_attachment=True,
            download_name=f'checklist_submissions_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    else:
        return jsonify({'error': 'No submissions file found'}), 404

@app.route('/health')
def health():
    """Health check endpoint"""
    status = {
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'email_configured': bool(ADMIN_EMAIL and SMTP_USERNAME and SMTP_PASSWORD),
        'google_sheets_configured': bool(GOOGLE_SHEETS_CREDS),
        'excel_file_exists': os.path.exists(SUBMISSIONS_FILE)
    }
    return jsonify(status)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print("="*50)
    print("Load Bank Checklist System Starting...")
    print("="*50)
    print(f"Port: {port}")
    print(f"Email configured: {bool(ADMIN_EMAIL and SMTP_USERNAME and SMTP_PASSWORD)}")
    print(f"Google Sheets configured: {bool(GOOGLE_SHEETS_CREDS)}")
    print("="*50)
    app.run(host='0.0.0.0', port=port, debug=False)
